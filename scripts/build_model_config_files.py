import sys
sys.path.append('../commonfiles/python')
import os
import logging.config
import csv
import glob
import optparse
if sys.version_info[0] < 3:
  import ConfigParser
else:
  import configparser as ConfigParser
import traceback
from openpyxl import load_workbook
from mb_wq_data import mb_sample_sites

VB_FUNCTIONS_REPLACE = [
  ("POLY", "VB_POLY"),
  ("QUADROOT", "VB_QUADROOT"),
  ("SQUAREROOT", "VB_SQUAREROOT"),
  ("INVERSE", "VB_INVERSE"),
  ("SQUARE", "VB_SQUARE"),
  ("WindO_comp", "VB_WindO_comp"),
  ("WindA_comp", "VB_WindA_comp"),
  ("LOG10", "VB_LOG10")
]

excel_sheet_columns = ["Site",
                       "# of Observations",
                       "# of Variables",
                       "Model for Log Enterococci",
                       "BIC",
                       "R2",
                       "Adj-R2",
                       "AUC"]

def cleanup_mlr(formula_string, location):
  no_log_10 = False
  if formula_string.find('etcoc =') != -1 or formula_string.find('enterococcus_value =') != -1:
    no_log_10 = True
  formula_string = formula_string.replace('[', '(')
  formula_string = formula_string.replace(']', ')')
  if location:
    formula_string = formula_string.replace('radar_rain_summary', '%s_nexrad_summary' % (location))
    formula_string = formula_string.replace('radar_rainfall_intensity_24', '%s_nexrad_rainfall_intensity' % (location))
    formula_string = formula_string.replace('radar_rainfall', '%s_nexrad_rainfall' % (location))
    formula_string = formula_string.replace('radar_preceding_dry_day_cnt', '%s_nexrad_dry_days_count' % (location))
    formula_string = formula_string.replace('radar_rain_total_one_day_delay', '%s_nexrad_total_1_day_delay' % (location))
    formula_string = formula_string.replace('radar_rain_total_two_day_delay', '%s_nexrad_total_2_day_delay' % (location))
    formula_string = formula_string.replace('radar_rain_total_three_day_delay',
                                            '%s_nexrad_total_3_day_delay' % (location))
    formula_string = formula_string.replace('range', 'tide_range_8661070')

  formula_string = formula_string.replace('apache_', 'apachepier_')
  if no_log_10:
    formula_string = formula_string.replace('etcoc = ', '')
    formula_string = formula_string.replace('enterococcus_value = ', '')
  else:
    formula_string = formula_string.replace('LOG10(etcoc) =', '')
    formula_string = formula_string.replace('LOG10(enterococcus_value) =', '')
    formula_string = 'Pow(10, %s)' % (formula_string)
  for vb_replacement in VB_FUNCTIONS_REPLACE:
    formula_string = formula_string.replace(vb_replacement[0], vb_replacement[1])

  return formula_string

def process_excel_file(**kwargs):
  logger = logging.getLogger(__name__)
  excel_file = kwargs['source_file']
  test_sites = kwargs['test_sites']
  watersheds = kwargs['watersheds']
  output_dir = kwargs['output_dir']
  wb = load_workbook(filename = excel_file)

  ws_sites = wb.get_sheet_names()
  try:
    for sample_site in test_sites:

      logger.info("Site: %s" % (sample_site.name))

      matches = []
      site_name = sample_site.name.replace('-', '').upper()
      #SOme sites have an A, ie WAC-005A. Need to distiguish a WAC-005 from WAC-005A site.
      a_site = False
      if site_name[-1] == 'A':
        a_site = True

      if site_name == 'WAC015A':
        site_name
      for s_name in ws_sites:
        if s_name.find(site_name) != -1:
          check_char = None
          if len(s_name) > len(site_name):
            if a_site:
              check_char = s_name[len(site_name)-1]
            else:
              check_char = s_name[len(site_name)]
          if check_char is not None:
            if not a_site and check_char != "A":
              matches.append(s_name)
            elif a_site and check_char == 'A':
              matches.append(s_name)
          else:
            matches.append(s_name)

      if len(matches):
        logger.info("Creating model config for: %s" % (sample_site.name))
        model_config_parser = ConfigParser.ConfigParser()
        model_config_outfile_name = os.path.join(output_dir, '%s.ini' % (sample_site.name.lower()))

        # Check to see if the model ini file exists.
        ini_exists = False
        if os.path.isfile(model_config_outfile_name):
          model_config_parser.read(model_config_outfile_name)
          ini_exists = True
          model_num = model_config_parser.getint("settings", "model_count") + 1

        if not ini_exists:
          model_config_parser.add_section("settings")

        model_num = 1
        with open(model_config_outfile_name, 'w') as model_config_ini_obj:
          for model_sheet in matches:
            logger.debug("Processing worksheet: %s" % (model_sheet))
            data_sheet = wb[model_sheet]
            for index, row in enumerate(data_sheet.iter_rows()):
              if index >= 1:
                formula_string = row[3].value
                formula_string = cleanup_mlr(formula_string, None)

                model_section = "model_%d" % model_num
                model_name = "%s-%d" % (model_sheet, index)
                model_config_parser.add_section(model_section)
                model_config_parser.set(model_section, 'name', model_name)
                model_config_parser.set(model_section, 'formula', formula_string)

                model_num += 1
          model_config_parser.set("settings", "model_count", model_num-1)
          model_config_parser.write(model_config_ini_obj)

  except Exception as e:
    logger.exception(e)
  return

def main():
  parser = optparse.OptionParser()
  parser.add_option("-c", "--ConfigFile", dest="config_file",
                    help="INI Configuration file." )
  parser.add_option("-d", "--DestinationDirectory", dest="dest_dir",
                    help="Destination directory to write the config files." )
  parser.add_option("-m", "--ModelCSVFileDirectory", dest="model_csv_dir", default="",
                    help="Directory where the CSV files defining the models are located." )
  parser.add_option("-x", "--ExcelModelFile", dest="excel_file", default="",
                    help="Excel file with the model equations." )

  (options, args) = parser.parse_args()

  logger = None

  if(options.config_file is None):
    parser.print_help()
    sys.exit(-1)

  try:
    config_file = ConfigParser.RawConfigParser()
    config_file.read(options.config_file)

    logger = None
    logConfFile = config_file.get('logging', 'prediction_engine')
    if logConfFile:
      logging.config.fileConfig(logConfFile)
      logger = logging.getLogger('mb_wq_predicition_logger')
      logger.info("Log file opened.")

  except ConfigParser.Error, e:
    traceback.print_exc(e)
    sys.exit(-1)

  try:
    boundaries_location_file = config_file.get('boundaries_settings', 'boundaries_file')
    sites_location_file = config_file.get('boundaries_settings', 'sample_sites')
  except ConfigParser.Error as e:
    if logger:
      logger.exception(e)
  else:
    #Load the sample site information. Has name, location and the boundaries that contain the site.
    mb_sites = mb_sample_sites()
    mb_sites.load_sites(file_name=sites_location_file, boundary_file=boundaries_location_file)
    #Build watershed groups.
    watersheds = {}
    for site in mb_sites:
      if site.contained_by[0] is not None:
        if site.contained_by[0].name.lower() not in watersheds:
          watersheds[site.contained_by[0].name.lower()] = []
        watersheds[site.contained_by[0].name.lower()].append(site.name)

    if len(options.model_csv_dir):
      model_csv_list = glob.glob('%s/*.csv' % (options.model_csv_dir))
      start_line = 1
      header_row = [
        "Location",
        "Site",
        "Equation"
      ]

      for model_csv_file in model_csv_list:
        with open(model_csv_file, "rU") as model_file_obj:
          path, filename_ext = os.path.split(model_csv_file)
          filename, ext = os.path.splitext(filename_ext)
          model_file_reader = csv.DictReader(model_file_obj, delimiter=',', quotechar='"', fieldnames=header_row)
          line_num = 0
          current_watershed = None
          sites = None
          sites_list = []
          for row in model_file_reader:
            #Actually model lines start multiple lines into the file.
            if line_num >= start_line:
              location = row['Location'].lower()
              if current_watershed is None or current_watershed != location:
                sites = watersheds[location]
              site = row['Site']
              if site in sites:
                sites.remove(site)
                sites_list.append(site)
              else:
                sites_list = sites
              for sample_site in sites_list:
                logger.info("Creating model config for: %s" % (sample_site))
                model_config_parser = ConfigParser.ConfigParser()
                model_config_outfile_name = os.path.join(options.dest_dir, '%s.ini' % (sample_site.lower()))
                try:
                  #Check to see if the model ini file exists.
                  ini_exists = False
                  model_num = 1
                  if os.path.isfile(model_config_outfile_name):
                    model_config_parser.read(model_config_outfile_name)
                    ini_exists = True
                    model_num = model_config_parser.getint("settings", "model_count") + 1
                  with open(model_config_outfile_name, 'w') as model_config_ini_obj:

                    if not ini_exists:
                      model_config_parser.add_section("settings")

                    formula_string = row['Equation']

                    no_log_10 = False
                    if formula_string.find('etcoc =') != -1:
                      no_log_10 = True
                    formula_string = formula_string.replace('[', '(')
                    formula_string = formula_string.replace(']', ')')

                    formula_string = formula_string.replace('radar_rain_summary', '%s_nexrad_summary' % (location))
                    formula_string = formula_string.replace('radar_rainfall_intensity_24', '%s_nexrad_rainfall_intensity' % (location))
                    formula_string = formula_string.replace('radar_rainfall', '%s_nexrad_rainfall' % (location))
                    formula_string = formula_string.replace('radar_preceding_dry_day_cnt', '%s_nexrad_dry_days_count' % (location))
                    formula_string = formula_string.replace('radar_rain_total_one_day_delay', '%s_nexrad_total_1_day_delay' % (location))
                    formula_string = formula_string.replace('radar_rain_total_two_day_delay', '%s_nexrad_total_2_day_delay' % (location))
                    formula_string = formula_string.replace('radar_rain_total_three_day_delay', '%s_nexrad_total_3_day_delay' % (location))
                    formula_string = formula_string.replace('range', 'tide_range_8661070')
                    if no_log_10:
                      formula_string = formula_string.replace('etcoc = ', '')
                    else:
                      formula_string = formula_string.replace('LOG10(etcoc) =', '')
                    formula_string = 'Pow(10, %s)' % (formula_string)
                    for vb_replacement in VB_FUNCTIONS_REPLACE:
                      formula_string = formula_string.replace(vb_replacement[0], vb_replacement[1])

                    model_section = "model_%d" % model_num
                    model_name = "%s-%s" % (sample_site, filename)
                    model_config_parser.add_section(model_section)
                    model_config_parser.set(model_section, 'name', model_name)
                    model_config_parser.set(model_section, 'formula', formula_string)
                    model_config_parser.set("settings", "model_count", model_num)
                    model_config_parser.write(model_config_ini_obj)

                except IOError, e:
                  if logger:
                    logger.exception(e)
              del sites_list[:]
            line_num += 1
    elif len(options.excel_file):
      process_excel_file(source_file=options.excel_file,
                         test_sites=mb_sites,
                         watersheds=watersheds,
                         output_dir=options.dest_dir)

  logger.info("Log file closed.")

  return


if __name__ == "__main__":
  main()
